"""XLSX reader backed by ``openpyxl``.

Strategy:
- Open the workbook read-only with ``openpyxl`` (``data_only=True`` so we
  get cell values, not formulas).
- Treat **each sheet as one ``Page``**. ``meta.pages`` therefore equals the
  sheet count; page numbers start at 1 in workbook order. Each ``Page``
  carries ``page.name = ws.title`` so consumers can address sheets by name
  (e.g. ``[p for p in doc.pages if p.name == "Summary"]``) — this keeps the
  surface uniform across formats (no separate ``Sheet`` model) while still
  exposing the sheet identity.
- Render every sheet as a single Markdown table via ``tabulate``. First row
  is the header (standard spreadsheet convention). Per-cell escaping is
  applied before handing off to ``tabulate`` (same three hazards as CSV:
  literal ``|``, embedded newlines, and number re-formatting that would
  drop leading zeros).
- Prefix each page's Markdown with an ``# {sheet_name}`` heading so the
  concatenated ``doc.text`` reads naturally.

File naming rule: ``openpyxl.py`` — the core driver. ``tabulate`` is a
post-processor and does not determine the file name.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from tabulate import tabulate

from ..base import Reader
from ..._core.document import Document, Meta, Page, Table
from ...errors import ParseError


def _escape_md_cell(cell: Any) -> str:
    """Escape a single cell so it is safe inside a Markdown pipe table."""
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        s = cell.isoformat()
    elif isinstance(cell, bytes):
        s = cell.decode("utf-8", errors="replace")
    else:
        s = str(cell)
    return s.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>").replace("\r", "<br>")


class XlsxReader(Reader):
    name = "openpyxl"
    formats = ("xlsx",)
    is_default = True

    def read(self, data: bytes, *, source_name: Optional[str] = None, **_) -> Document:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ParseError(
                "openpyxl is required for XLSX parsing: pip install openpyxl"
            ) from e

        warnings: list[str] = []

        try:
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception as e:
            raise ParseError(f"failed to open xlsx: {e}") from e

        try:
            pages: list[Page] = []
            for page_no, ws in enumerate(wb.worksheets, start=1):
                page = _render_sheet(ws, page_no, warnings)
                pages.append(page)

            if not pages:
                pages.append(Page(text="", number=1))
                warnings.append("workbook has no sheets")

            props = wb.properties
            title = (getattr(props, "title", None) or None) or None
            author = (getattr(props, "creator", None) or None) or None
            created_at = getattr(props, "created", None)
            if created_at is not None and created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if not title and source_name:
            title = Path(source_name).stem or None

        try:
            meta = Meta(
                format="xlsx",
                pages=len(pages),
                size=len(data),
                title=title,
                author=author,
                created_at=created_at,
                reader=self.name,
                warnings=warnings,
            )
            return Document(pages=pages, meta=meta)
        except Exception as e:  # pragma: no cover - defensive
            raise ParseError(f"xlsx reader failed to build Document: {e}") from e


def _render_sheet(ws, page_no: int, warnings: list[str]) -> Page:
    """Render one worksheet as a ``Page`` with a single Markdown table.

    Always sets ``page.name = ws.title`` so downstream code can identify
    sheets by their human name without having to look at ``page.text``.
    """
    sheet_name = ws.title
    sheet_heading = f"# {sheet_name}"

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        # Empty sheet.
        return Page(
            text=f"{sheet_heading}\n\n_(empty sheet)_",
            number=page_no,
            name=sheet_name,
        )

    headers = [_stringify(h) for h in header_row]
    body: list[list[str]] = []
    for row in rows_iter:
        body.append([_stringify(c) for c in row])

    safe_headers = [_escape_md_cell(h) for h in headers]
    safe_body = [[_escape_md_cell(c) for c in row] for row in body]
    md_table = tabulate(
        safe_body,
        headers=safe_headers,
        tablefmt="pipe",
        disable_numparse=True,
    )

    page_md = f"{sheet_heading}\n\n{md_table}" if md_table else f"{sheet_heading}\n\n_(no rows)_"
    table = Table(text=md_table, rows=body, headers=headers, page=page_no)
    return Page(text=page_md, number=page_no, name=sheet_name, tables=[table])


def _stringify(cell: Any) -> str:
    """Normalise an openpyxl cell value into a plain string for ``Table.rows``."""
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.isoformat()
    if isinstance(cell, bytes):
        return cell.decode("utf-8", errors="replace")
    return str(cell)
